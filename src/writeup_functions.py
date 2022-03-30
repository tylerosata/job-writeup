#!/usr/bin/python

from exchangelib import DELEGATE, Configuration, Credentials, Account, FileAttachment
import os
import PyPDF2
import re
import openpyxl
import glob
from datetime import datetime, date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import sys
import worklogin

os.environ['GH_TOKEN'] = worklogin.token
files = glob.glob('C:\\Users\\otyle\\spam\\new jobs\\*.pdf')
i = 0
bags = dict()
today = str(date.today())
now = int(today[8:10])
dueDate = now + 7
mainPanel = '_M'
combined = '_C'
gusset = '_G'
projectTitle = " ".join(map(str,sys.argv[1:]))
projectName = " ".join(map(str, sys.arg[1:3]))
browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

def hubxlogin(): # should move login to access.py

    browser.get('https://access.sgkinc.com')
    userElem = browser.find_element(By.NAME,'usr')
    userElem.send_keys(worklogin.username)
    userPwd = browser.find_element(By.NAME,'pwd')
    userPwd.send_keys(worklogin.pwd)
    loginButton = browser.find_element(By.CLASS_NAME,'btn')
    loginButton.click()

def createProject():
    
    projects = browser.find_element(By.XPATH, "//*[@id='pgMenu']/tbody/tr/td[2]")
    projects.click()
    newProject = browser.find_element(By.XPATH, "//*[@id='pgCont']//button[1]")
    newProject.click()
    projectName = browser.find_element(By.XPATH, "//*[@id='in17']")
    projectName.send_keys(projectTitle + " " + str(today))
    selectSite = browser.find_element(By.ID, 'in20')
    select_site = Select(selectSite)
    select_site.select_by_value('5077')
    parentGroup = browser.find_element(By.ID, 'in12')
    select_parent = Select(parentGroup)
    select_parent.select_by_value('1000000788')
    time.sleep(2)
    soldTo = browser.find_element(By.ID, 'in21')
    select_soldTo = Select(soldTo)
    select_soldTo.select_by_visible_text('BLUE BUFFALO COMPANY LTD | 1000024772')
    time.sleep(2)
    csr = browser.find_element(By.ID, 'in23')
    select_csr = Select(csr)
    select_csr.select_by_value('2817')
    time.sleep(2)
    zfoc = browser.find_element(By.ID, 'in11')
    select_zfoc = Select(zfoc)
    select_zfoc.select_by_value('ZFOC')
    save = browser.find_element(By.XPATH, "//*[@id='d1']/button[1]")
    save.click()
    time.sleep(2)

def projAddFisrtJob():
    
    ###switch to jobs tab in project and create first job
    browser.find_element(By.ID, "tabsub").click()
    time.sleep(2)
    browser.find_element(By.XPATH, '//*[@id="pgCont"]/div/table[2]/thead/tr[2]/td/table/tbody/tr/td[1]/button').click()
    time.sleep(2)
    
def soldToNoProj(): #adds parent group and sold to for jobs created individually and not within a project
    
    parentGroup = browser.find_element(By.ID, 'in22')
    selectParent = Select(parentGroup)
    selectParent.select_by_value('1000000788')
    time.sleep(2)
    soldTo = browser.find_element(By.ID, 'in30')
    selectBB = Select(soldTo)
    selectBB.select_by_value('1000024772')
    
def serviceOrder(MTnumber, packSize, dieline, description, UPC, panel):
    
    if "BB" not in dieline:
        panel = mainPanel
    else:
        panel = combined
        
    # swith to TECHNICAL tab
    browser.find_element(By.XPATH, '//*[@id="tabdet"]').click()
    select_UPCA = browser.find_element(By.ID, 'in155')
    selectUPC = Select(select_UPCA)
    selectUPC.select_by_value('038')
    UPCinput = browser.find_element(By.ID, 'in143')
    UPCinput.send_keys(UPC) # input UPC number
    browser.find_element(By.ID, 'in613').send_keys(projectName) # PROJECT NAME

    #Switch back to GENERAL TAB
    browser.find_element(By.ID, 'tabjob').click()
    browser.find_element(By.ID, 'in24').send_keys('TBD')  # PO NUMBER
    time.sleep(2)
    selectPrinter = browser.find_element(By.ID, 'in25')
    choosePrinter = Select(selectPrinter)
    choosePrinter.select_by_visible_text('Amcor Flexibles Oshkosh North | 1000002470')
    category = browser.find_element(By.ID, 'in34')
    chooseCat = Select(category)
    chooseCat.select_by_value('Dry')
    variantInput = browser.find_element(By.ID, 'in44')
    variantInput.send_keys('MT' + str(MTnumber) + panel) #input variant
    bagName1 = browser.find_element(By.ID, 'in41')
    bagName1.send_keys(description + panel) # description / name
    browser.find_element(By.ID, 'in39').send_keys('Packaging') # PACK TYPE
    brand = browser.find_element(By.ID, 'in33')
    chooseBrand = Select(brand)
    chooseBrand.select_by_value('Blue Buffalo')
    bagName2 = browser.find_element(By.ID, 'in36')
    bagName2.send_keys(description + panel) # MATERIAL DESCRIPTION
    packageSize = browser.find_element(By.ID, 'in38')
    packageSize.send_keys(packSize) # PACK SIZE
    printMethod = browser.find_element(By.ID, 'in56')
    choosePrintMethod = Select(printMethod)
    choosePrintMethod.select_by_value('03') # FLEXO
    printSpec = browser.find_element(By.ID, 'in57')
    choosePrintSpec = Select(printSpec)
    time.sleep(1)
    choosePrintSpec.select_by_value('P1802073989') # For AMCOR NORTH this is 1718
    dielineInput1 = browser.find_element(By.ID, 'in46')
    dielineInput1.send_keys(dieline) # first dieline input
    dielineInput2 = browser.find_element(By.ID, 'in45')
    dielineInput2.send_keys(dieline) # second dieline input
    inputBMN = browser.find_element(By.ID, 'in72')
    inputBMN.send_keys('MT' + str(MTnumber)) #input BMN field
    csr = browser.find_element(By.ID, 'in88')
    chooseCSR = Select(csr)
    chooseCSR.select_by_value('2817') #Select CSR - may not be needed when started from project

def datepicker():

    # first calendar date picker
    browser.find_element(By.XPATH, '//*[@id="frm.ddl"]/div/table[1]/tbody/tr[1]/td[2]/table/tbody/tr/td[1]/table/tbody/tr/td[3]/label/img').click() 
    # gets to the actual calendar part (days and weeks)
    dates = browser.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody')
    last = dates.find_elements(By.CLASS_NAME, 'ui-state-default')[-1]
    lastDay = int(last.text)
    # this works to select the next month
    nextMonth = browser.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[2]')
    if dueDate > lastDay:
        newDue = dueDate - lastDay
        nextMonth.click()
        newCalendar = browser.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody')
        newCalendar.find_element(By.LINK_TEXT, str(newDue)).click()
    else:
        days = dates.find_element(By.LINK_TEXT, str(dueDate))
        days.click()
        #this is the second date picker, clicks two months forward and selects the last date - weekend or weekday doesn't matter
    browser.find_element(By.XPATH, '//*[@id="frm.ddl"]/div/table[1]/tbody/tr[1]/td[4]/table/tbody/tr/td[1]/table/tbody/tr/td[3]/label/img').click()
    #the below commented  out code is a repeat of above in the function so do I need it again? yes because selenium
    dates = browser.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody')
    dates = browser.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody')
    last = dates.find_elements(By.CLASS_NAME, 'ui-state-default')[-1]
    last.click() 
    
def ZFOC(): #this is to make them no charge jobs so I can test save and dupe
          
    orderType = browser.find_element(By.ID,'in21')
    chooseOrderType = Select(orderType)
    chooseOrderType.select_by_value('ZFOC')
    
def savejob():

    savePanel = browser.find_element(By.CLASS_NAME, 'jobpan')
    save = savePanel.find_elements(By.TAG_NAME, 'button')[0]
    save.click()
    time.sleep(5)

def copyInProject(): # can I manage to open in new tab and move to that tab to continue? not necessary but would be convenient

    browser.find_element(By.XPATH, '//*[@id="l40"]').click() #opens the copy options
    time.sleep(1)
    browser.find_element(By.XPATH, '//*[@id="d39"]/table/tbody/tr[1]/td[2]').click() #clicks copy in project
    # browser.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/form/table/tbody/tr/td[3]/div/div/[2]/div[2]/button[2]/div/table/tbody/tr/td[2]').click()

def gusset(MTnumber, description): #adds gusset after Main panel

    variantInput = browser.find_element(By.ID, 'in44')
    variantInput.clear()
    variantInput.send_keys('MT' + str(MTnumber) + '_G') # variant
    bagName1 = browser.find_element(By.ID, 'in41')
    bagName1.clear()
    bagName1.send_keys(description + '_G') # description 1
    bagName2 = browser.find_element(By.ID, 'in36')
    bagName2.clear()
    bagName2.send_keys(description + '_G') # description 2

def newBag(MTnumber, packSize, dieline, description, UPC, panel): #either a new Main panel or combined

    if "BB" not in dieline:
        panel = mainPanel
    else:
        panel = combined
        
    browser.find_element(By.XPATH, '//*[@id="tabdet"]').click() # swith to TECHNICAL tab
    UPCinput = browser.find_element(By.ID, 'in143')
    UPCinput.clear()
    UPCinput.send_keys(UPC) # input UPC number
        
    browser.find_element(By.ID, 'tabjob').click() #Switch back to GENERAL TAB
    variantInput = browser.find_element(By.ID, 'in44')
    variantInput.clear()
    variantInput.send_keys('MT' + str(MTnumber) + panel)
    bagName1 = browser.find_element(By.ID, 'in41')
    bagName1.clear()
    bagName1.send_keys(description + panel)
    bagName2 = browser.find_element(By.ID, 'in36')
    bagName2.clear()
    bagName2.send_keys(description + panel)
    packageSize = browser.find_element(By.ID, 'in38')
    packageSize.clear()
    packageSize.send_keys(packSize) # PACXK SIZE
    dielineInput1 = browser.find_element(By.ID, 'in46')
    dielineInput1.clear()
    dielineInput1.send_keys(dieline) # first dieline input
    dielineInput2 = browser.find_element(By.ID, 'in45')
    dielineInput2.clear()
    dielineInput2.send_keys(dieline) # second dieline input
    inputBMN = browser.find_element(By.ID, 'in72')
    inputBMN.clear()
    inputBMN.send_keys('MT' + str(MTnumber)) #input BMN field
    
def firstBag():
    if "BB" not in details[2]:
        panel = mainPanel
        serviceOrder(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        copyInProject()
        gusset(details[0],details[3])
        savejob()
    else:
        panel = combined
        serviceOrder(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        
def nextBags():
    if "BB" not in details[2]:
        panel = mainPanel
        copyInProject()
        newBag(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        copyInProject()
        gusset(details[0],details[3])
        savejob()
    else:
        panel = combined
        copyInProject()
        newBag(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()

##def main():
##    
##    hubxlogin()
##    createProject()
##    addJob()
##
##    for i in bags:
##        details = bags.get(i)
##        if i <= 0:
##            firstBag()
##        else:
##            nextBags()

while i < len(files):
        # gets the mt number and bag size from pdf name
    MTnumber = int(re.search(r'\d{6}',files[i]).group())
    packSize = re.search(r'\d?\.?\d?#',files[i]).group()
    print(MTnumber)
    print(packSize)
        #get the dieline, UPC, and description(name) from tracker based on MT number from pdf
    wb = openpyxl.load_workbook('timeline_tracker.xlsx')
    sheet = wb['Timeline']
    for rowNum in range(4, sheet.max_row):
        mtNumbers = sheet.cell(row=rowNum, column = 17).value
        if MTnumber == mtNumbers:
            dieline = sheet.cell(row=rowNum, column=13).value
            description = sheet.cell(row=rowNum, column=12).value 
            UPC = sheet.cell(row=rowNum, column=11).value
    wb.close()
    bags[i] = [MTnumber, packSize, dieline, description, UPC]
    i +=1
    print(bags)
##
##if __name__=="__main__":
##    main()

hubxlogin()
createProject()
projAddFisrtJob()

for i in bags:
    details = bags.get(i)
    if i <= 0:
        firstBag()
    else:
        nextBags()
